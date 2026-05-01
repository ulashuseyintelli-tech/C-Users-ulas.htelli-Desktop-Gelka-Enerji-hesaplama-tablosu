"""
Unit tests for backend/app/pricing/excel_parser.py

Tests cover:
- expected_hours_for_period() — month-day calculation
- parse_epias_excel() — EPİAŞ uzlaştırma Excel parsing
- parse_consumption_excel() — customer consumption Excel parsing
- Edge cases: invalid format, missing columns, out-of-range values, duplicates
"""

from __future__ import annotations

import calendar
import io
from datetime import datetime

import pytest
from openpyxl import Workbook

from app.pricing.excel_parser import (
    EpiasParseOutput,
    ConsumptionParseOutput,
    ParsedMarketRecord,
    expected_hours_for_period,
    parse_epias_excel,
    parse_consumption_excel,
)


# ═══════════════════════════════════════════════════════════════════════════════
# Helpers — Excel dosyası oluşturma
# ═══════════════════════════════════════════════════════════════════════════════


def _make_epias_excel(
    year: int = 2026,
    month: int = 3,
    region: str = "TR1",
    ptf_base: float = 1500.0,
    smf_base: float = 1600.0,
    sheet_name: str = "Uzlaştırma Dönemi Detayı",
    skip_hours: list[int] | None = None,
    duplicate_hours: list[int] | None = None,
    extra_rows: list[tuple] | None = None,
) -> bytes:
    """EPİAŞ formatında test Excel dosyası oluştur.

    Gerçek format: Tarih (datetime), Versiyon (datetime), Bölge, PTF, SMF
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header
    ws.append(["Tarih", "Versiyon", "Bölge", "PTF (TL / MWh)", "SMF (TL / MWh)"])

    days_in_month = calendar.monthrange(year, month)[1]
    skip_set = set(skip_hours or [])
    dup_set = set(duplicate_hours or [])

    for day in range(1, days_in_month + 1):
        for hour in range(24):
            idx = (day - 1) * 24 + hour
            if idx in skip_set:
                continue

            dt = datetime(year, month, day, hour, 0)
            version_dt = datetime(year, month, 1, 0, 0)
            ptf = ptf_base + (hour * 10)  # Vary by hour
            smf = smf_base + (hour * 12)

            ws.append([dt, version_dt, region, ptf, smf])

            # Duplicate rows
            if idx in dup_set:
                ws.append([dt, version_dt, region, ptf + 1, smf + 1])

    # Extra rows (for testing invalid data)
    if extra_rows:
        for row_data in extra_rows:
            ws.append(list(row_data))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_consumption_excel(
    year: int = 2026,
    month: int = 3,
    with_saat_column: bool = True,
    negative_hours: list[int] | None = None,
) -> bytes:
    """Tüketim formatında test Excel dosyası oluştur."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Tüketim"

    neg_set = set(negative_hours or [])
    days_in_month = calendar.monthrange(year, month)[1]

    if with_saat_column:
        ws.append(["Tarih", "Saat", "Tüketim (kWh)"])
        for day in range(1, days_in_month + 1):
            for hour in range(24):
                dt = datetime(year, month, day)
                idx = (day - 1) * 24 + hour
                kwh = -50.0 if idx in neg_set else 100.0 + hour
                ws.append([dt, hour, kwh])
    else:
        # EPİAŞ-like format: datetime with hour embedded
        ws.append(["Tarih", "Tüketim (kWh)"])
        for day in range(1, days_in_month + 1):
            for hour in range(24):
                dt = datetime(year, month, day, hour, 0)
                idx = (day - 1) * 24 + hour
                kwh = -50.0 if idx in neg_set else 100.0 + hour
                ws.append([dt, kwh])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════════
# Tests: expected_hours_for_period
# ═══════════════════════════════════════════════════════════════════════════════


class TestExpectedHoursForPeriod:
    """expected_hours_for_period() testleri."""

    def test_31_day_month(self):
        assert expected_hours_for_period("2026-03") == 744

    def test_30_day_month(self):
        assert expected_hours_for_period("2026-04") == 720

    def test_28_day_february(self):
        assert expected_hours_for_period("2025-02") == 672

    def test_29_day_leap_february(self):
        assert expected_hours_for_period("2024-02") == 696

    def test_january(self):
        assert expected_hours_for_period("2025-01") == 744

    def test_invalid_format_raises(self):
        with pytest.raises(ValueError, match="Geçersiz dönem formatı"):
            expected_hours_for_period("2025-13")

    def test_invalid_string_raises(self):
        with pytest.raises(ValueError, match="Geçersiz dönem formatı"):
            expected_hours_for_period("not-a-period")

    def test_empty_string_raises(self):
        with pytest.raises(ValueError, match="Geçersiz dönem formatı"):
            expected_hours_for_period("")

    def test_month_00_raises(self):
        with pytest.raises(ValueError, match="Geçersiz dönem formatı"):
            expected_hours_for_period("2025-00")


# ═══════════════════════════════════════════════════════════════════════════════
# Tests: parse_epias_excel
# ═══════════════════════════════════════════════════════════════════════════════


class TestParseEpiasExcel:
    """parse_epias_excel() testleri."""

    def test_valid_full_month(self):
        """Tam bir ay verisi — 744 saat, hatasız."""
        data = _make_epias_excel(year=2026, month=3)
        output = parse_epias_excel(data, "test.xlsx")

        assert output.result.success is True
        assert output.result.period == "2026-03"
        assert output.result.total_rows == 744
        assert output.result.expected_hours == 744
        assert output.result.missing_hours == []
        assert output.result.rejected_rows == []
        assert output.result.quality_score == 100
        assert len(output.records) == 744

    def test_period_detection(self):
        """Dönem doğru çıkarılıyor."""
        data = _make_epias_excel(year=2025, month=1)
        output = parse_epias_excel(data, "test.xlsx")

        assert output.result.period == "2025-01"
        assert output.result.expected_hours == 744

    def test_30_day_month(self):
        """30 günlük ay — 720 saat."""
        data = _make_epias_excel(year=2026, month=4)
        output = parse_epias_excel(data, "test.xlsx")

        assert output.result.total_rows == 720
        assert output.result.expected_hours == 720

    def test_missing_hours_detected(self):
        """Eksik saatler tespit ediliyor."""
        data = _make_epias_excel(year=2026, month=3, skip_hours=[0, 1, 100])
        output = parse_epias_excel(data, "test.xlsx")

        assert output.result.success is True
        assert output.result.total_rows == 741
        assert len(output.result.missing_hours) == 3
        assert 0 in output.result.missing_hours
        assert 1 in output.result.missing_hours
        assert 100 in output.result.missing_hours
        assert any("eksik saat" in w.lower() for w in output.result.warnings)

    def test_duplicate_hours_detected(self):
        """Mükerrer saatler tespit ediliyor ve kaldırılıyor."""
        data = _make_epias_excel(year=2026, month=3, duplicate_hours=[0, 50])
        output = parse_epias_excel(data, "test.xlsx")

        assert output.result.success is True
        assert output.result.total_rows == 744  # Duplicates removed
        assert any("mükerrer" in w.lower() for w in output.result.warnings)

    def test_region_filter(self):
        """TR1 olmayan bölgeler filtreleniyor."""
        # Non-TR1 region → all rows filtered out
        data = _make_epias_excel(year=2026, month=3, region="TR2")
        output = parse_epias_excel(data, "test.xlsx")

        # Period is still detected from datetime, but no valid rows
        assert output.result.total_rows == 0
        assert len(output.records) == 0
        assert len(output.result.missing_hours) == 744

    def test_ptf_out_of_range_rejected(self):
        """PTF aralık dışı değerler reddediliyor."""
        extra = [(datetime(2026, 3, 1, 0, 0), datetime(2026, 3, 1), "TR1", -100, 1500)]
        data = _make_epias_excel(
            year=2026, month=3,
            skip_hours=[0],  # Skip hour 0 so we can add invalid one
            extra_rows=extra,
        )
        output = parse_epias_excel(data, "test.xlsx")

        assert output.result.success is True
        assert any("PTF aralık dışı" in r["reason"] for r in output.result.rejected_rows)

    def test_invalid_excel_format(self):
        """Geçersiz Excel dosyası — hata döner."""
        output = parse_epias_excel(b"not an excel file", "bad.xlsx")

        assert output.result.success is False
        assert output.result.quality_score == 0

    def test_no_header_found(self):
        """Header bulunamayan Excel — hata döner."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Foo", "Bar", "Baz"])
        ws.append([1, 2, 3])
        buf = io.BytesIO()
        wb.save(buf)

        output = parse_epias_excel(buf.getvalue(), "no_header.xlsx")
        assert output.result.success is False

    def test_fallback_to_first_sheet(self):
        """Uzlaştırma sheet'i yoksa ilk sheet kullanılır."""
        data = _make_epias_excel(sheet_name="Veri")
        output = parse_epias_excel(data, "test.xlsx")

        assert output.result.success is True
        assert any("ilk sheet" in w.lower() for w in output.result.warnings)

    def test_records_have_correct_fields(self):
        """Kayıtlar doğru alanları içeriyor."""
        data = _make_epias_excel(year=2026, month=3)
        output = parse_epias_excel(data, "test.xlsx")

        rec = output.records[0]
        assert rec.period == "2026-03"
        assert rec.date == "2026-03-01"
        assert rec.hour == 0
        assert isinstance(rec.ptf_tl_per_mwh, float)
        assert isinstance(rec.smf_tl_per_mwh, float)

    def test_quality_score_decreases_with_issues(self):
        """Sorunlar kalite skorunu düşürüyor."""
        # Full month — perfect score
        data_full = _make_epias_excel(year=2026, month=3)
        out_full = parse_epias_excel(data_full, "test.xlsx")

        # Missing hours — lower score
        data_missing = _make_epias_excel(year=2026, month=3, skip_hours=list(range(50)))
        out_missing = parse_epias_excel(data_missing, "test.xlsx")

        assert out_full.result.quality_score > out_missing.result.quality_score

    def test_smf_column_missing_warning(self):
        """SMF sütunu yoksa uyarı üretilir ve SMF=0 olur."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Uzlaştırma Dönemi Detayı"
        ws.append(["Tarih", "Bölge", "PTF (TL / MWh)"])

        days = calendar.monthrange(2026, 3)[1]
        for day in range(1, days + 1):
            for hour in range(24):
                dt = datetime(2026, 3, day, hour, 0)
                ws.append([dt, "TR1", 1500.0])

        buf = io.BytesIO()
        wb.save(buf)

        output = parse_epias_excel(buf.getvalue(), "no_smf.xlsx")
        assert output.result.success is True
        assert any("SMF sütunu bulunamadı" in w for w in output.result.warnings)
        # All SMF values should be 0.0
        assert all(r.smf_tl_per_mwh == 0.0 for r in output.records)


# ═══════════════════════════════════════════════════════════════════════════════
# Tests: parse_consumption_excel
# ═══════════════════════════════════════════════════════════════════════════════


class TestParseConsumptionExcel:
    """parse_consumption_excel() testleri."""

    def test_valid_with_saat_column(self):
        """Tarih + Saat + Tüketim formatı — tam ay."""
        data = _make_consumption_excel(year=2026, month=3, with_saat_column=True)
        output = parse_consumption_excel(data, "tuketim.xlsx", "CUST-001")

        assert output.result.success is True
        assert output.result.customer_id == "CUST-001"
        assert output.result.period == "2026-03"
        assert output.result.total_rows == 744
        assert output.result.total_kwh > 0
        assert output.result.negative_hours == []
        assert len(output.records) == 744

    def test_valid_without_saat_column(self):
        """Tarih (datetime gömülü saat) + Tüketim formatı."""
        data = _make_consumption_excel(year=2026, month=3, with_saat_column=False)
        output = parse_consumption_excel(data, "tuketim.xlsx", "CUST-002")

        assert output.result.success is True
        assert output.result.total_rows == 744

    def test_negative_consumption_warning(self):
        """Negatif tüketim uyarı üretir ama reddetmez."""
        data = _make_consumption_excel(
            year=2026, month=3,
            with_saat_column=True,
            negative_hours=[0, 10, 50],
        )
        output = parse_consumption_excel(data, "tuketim.xlsx", "CUST-003")

        assert output.result.success is True
        assert len(output.result.negative_hours) == 3
        assert any("negatif" in w.lower() for w in output.result.warnings)
        # Total rows still includes negative hours
        assert output.result.total_rows == 744

    def test_invalid_excel(self):
        """Geçersiz dosya — hata döner."""
        output = parse_consumption_excel(b"bad data", "bad.xlsx", "CUST-X")
        assert output.result.success is False

    def test_no_header(self):
        """Header bulunamayan dosya — hata döner."""
        wb = Workbook()
        ws = wb.active
        ws.append(["X", "Y", "Z"])
        buf = io.BytesIO()
        wb.save(buf)

        output = parse_consumption_excel(buf.getvalue(), "no_header.xlsx", "CUST-X")
        assert output.result.success is False

    def test_quality_score_perfect(self):
        """Tam veri — yüksek kalite skoru."""
        data = _make_consumption_excel(year=2026, month=3)
        output = parse_consumption_excel(data, "tuketim.xlsx", "CUST-001")

        assert output.result.quality_score == 100

    def test_quality_score_with_negatives(self):
        """Negatif tüketim kalite skorunu düşürür."""
        data_clean = _make_consumption_excel(year=2026, month=3)
        out_clean = parse_consumption_excel(data_clean, "t.xlsx", "C1")

        data_neg = _make_consumption_excel(
            year=2026, month=3, negative_hours=list(range(100)),
        )
        out_neg = parse_consumption_excel(data_neg, "t.xlsx", "C2")

        assert out_clean.result.quality_score > out_neg.result.quality_score
