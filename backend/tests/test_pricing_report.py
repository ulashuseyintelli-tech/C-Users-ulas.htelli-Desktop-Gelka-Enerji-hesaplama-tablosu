"""
Pricing Risk Engine — Rapor Üretimi Testleri.

Task 22.1: PDF rapor üretimi (template render testi)
Task 22.2: Excel rapor üretimi
"""

import pytest
from app.pricing.pricing_report import generate_excel_report, _build_template_context


# ═══════════════════════════════════════════════════════════════════════════════
# Test Data
# ═══════════════════════════════════════════════════════════════════════════════

def _sample_analysis_result():
    """Örnek analiz sonucu — rapor testleri için."""
    return {
        "status": "ok",
        "period": "2026-01",
        "customer_id": "CUST-001",
        "weighted_prices": {
            "weighted_ptf_tl_per_mwh": 3027.04,
            "weighted_smf_tl_per_mwh": 3100.82,
            "arithmetic_avg_ptf": 2980.67,
            "arithmetic_avg_smf": 3050.0,
            "total_consumption_kwh": 100000.0,
            "total_cost_tl": 302703.98,
            "hours_count": 744,
        },
        "supplier_cost": {
            "weighted_ptf_tl_per_mwh": 3027.04,
            "yekdem_tl_per_mwh": 162.73,
            "imbalance_tl_per_mwh": 2.50,
            "total_cost_tl_per_mwh": 3192.27,
        },
        "pricing": {
            "multiplier": 1.05,
            "sales_price_tl_per_mwh": 3349.25,
            "gross_margin_tl_per_mwh": 156.98,
            "dealer_commission_tl_per_mwh": 3.14,
            "net_margin_tl_per_mwh": 153.84,
            "total_sales_tl": 334925.87,
            "total_cost_tl": 318976.98,
            "total_gross_margin_tl": 15948.88,
            "total_dealer_commission_tl": 318.98,
            "total_net_margin_tl": 15379.91,
        },
        "time_zone_breakdown": {
            "T1": {
                "label": "Gündüz (06:00-16:59)",
                "consumption_kwh": 50000.0,
                "consumption_pct": 50.0,
                "weighted_ptf_tl_per_mwh": 2920.15,
                "weighted_smf_tl_per_mwh": 2980.40,
                "total_cost_tl": 160000.0,
            },
            "T2": {
                "label": "Puant (17:00-21:59)",
                "consumption_kwh": 25000.0,
                "consumption_pct": 25.0,
                "weighted_ptf_tl_per_mwh": 3400.80,
                "weighted_smf_tl_per_mwh": 3460.50,
                "total_cost_tl": 90000.0,
            },
            "T3": {
                "label": "Gece (22:00-05:59)",
                "consumption_kwh": 25000.0,
                "consumption_pct": 25.0,
                "weighted_ptf_tl_per_mwh": 2498.30,
                "weighted_smf_tl_per_mwh": 2545.60,
                "total_cost_tl": 52703.98,
            },
        },
        "loss_map": {
            "total_loss_hours": 218,
            "total_loss_tl": -14484.87,
            "by_time_zone": {"T1": 50, "T2": 120, "T3": 48},
            "worst_hours": [
                {"date": "2026-01-15", "hour": 18, "ptf": 4250.0, "sales_price": 3890.5, "loss_tl": -45.2},
                {"date": "2026-01-20", "hour": 19, "ptf": 4100.0, "sales_price": 3890.5, "loss_tl": -38.1},
            ],
        },
        "risk_score": {
            "score": "Düşük",
            "weighted_ptf": 3027.04,
            "arithmetic_avg_ptf": 2980.67,
            "deviation_pct": 1.56,
            "t2_consumption_pct": 25.0,
            "peak_concentration": 28.5,
            "reasons": ["Ağırlıklı PTF sapması düşük (%1.6)"],
        },
        "safe_multiplier": {
            "safe_multiplier": 1.100,
            "recommended_multiplier": 1.10,
            "confidence_level": 0.95,
            "periods_analyzed": 1,
            "monthly_margins": [31009.78],
            "warning": "Bu profil için ×1.10 altında güvenli katsayı bulunamadı.",
        },
        "warnings": [
            {"type": "coherence_warning", "message": "Risk düşük ama katsayı yüksek."},
        ],
        "simulation": [
            {"multiplier": 1.02, "total_sales_tl": 325356.55, "total_cost_tl": 318976.98,
             "gross_margin_tl": 6379.57, "dealer_commission_tl": 127.59, "net_margin_tl": 6001.98,
             "loss_hours": 292, "total_loss_tl": -18036.82},
            {"multiplier": 1.05, "total_sales_tl": 334925.87, "total_cost_tl": 318976.98,
             "gross_margin_tl": 15948.88, "dealer_commission_tl": 318.98, "net_margin_tl": 15379.91,
             "loss_hours": 218, "total_loss_tl": -14484.87},
            {"multiplier": 1.10, "total_sales_tl": 350874.72, "total_cost_tl": 318976.98,
             "gross_margin_tl": 31897.74, "dealer_commission_tl": 637.95, "net_margin_tl": 31009.78,
             "loss_hours": 155, "total_loss_tl": -10425.49},
        ],
        "hour_costs": [
            {"date": "2026-01-01", "hour": 0, "consumption_kwh": 120.5,
             "ptf_tl_per_mwh": 2200.0, "smf_tl_per_mwh": 2250.0,
             "yekdem_tl_per_mwh": 162.73, "base_cost_tl": 284.73,
             "sales_price_tl": 402.15, "margin_tl": 117.42,
             "is_loss_hour": False, "time_zone": "T3"},
        ],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Template Context Testleri
# ═══════════════════════════════════════════════════════════════════════════════

class TestTemplateContext:
    """_build_template_context() testleri."""

    def test_basic_context(self):
        """Temel context alanları doldurulur."""
        result = _sample_analysis_result()
        ctx = _build_template_context(result, "Test Müşteri", "Ali Veli")

        assert ctx["period"] == "2026-01"
        assert ctx["customer_name"] == "Test Müşteri"
        assert ctx["weighted_ptf"] == 3027.04
        assert ctx["risk_level"] == "Düşük"
        assert ctx["safe_multiplier"] == 1.100
        assert len(ctx["risk_reasons"]) >= 1

    def test_coherence_note_extracted(self):
        """Tutarlılık uyarısı warnings'den çıkarılır."""
        result = _sample_analysis_result()
        ctx = _build_template_context(result, None, None)

        assert ctx["coherence_note"] is not None
        assert "katsayı" in ctx["coherence_note"].lower()

    def test_simulation_passed_through(self):
        """Simülasyon verisi context'e aktarılır."""
        result = _sample_analysis_result()
        ctx = _build_template_context(result, None, None)

        assert len(ctx["simulation"]) == 3


# ═══════════════════════════════════════════════════════════════════════════════
# Excel Rapor Testleri
# ═══════════════════════════════════════════════════════════════════════════════

class TestExcelReport:
    """generate_excel_report() testleri."""

    def test_generates_valid_xlsx(self):
        """Geçerli .xlsx dosyası üretilir."""
        result = _sample_analysis_result()
        excel_bytes = generate_excel_report(result, "Test Müşteri")

        assert len(excel_bytes) > 0
        # XLSX magic bytes: PK (ZIP format)
        assert excel_bytes[:2] == b'PK'

    def test_has_5_sheets(self):
        """5 sheet oluşturulur."""
        from openpyxl import load_workbook
        import io

        result = _sample_analysis_result()
        excel_bytes = generate_excel_report(result)

        wb = load_workbook(io.BytesIO(excel_bytes))
        assert len(wb.sheetnames) == 5
        assert "Özet" in wb.sheetnames
        assert "T1-T2-T3 Dağılım" in wb.sheetnames
        assert "Simülasyon" in wb.sheetnames
        assert "Saatlik Detay" in wb.sheetnames
        assert "Zarar Haritası" in wb.sheetnames

    def test_summary_sheet_has_data(self):
        """Özet sheet'inde veriler var."""
        from openpyxl import load_workbook
        import io

        result = _sample_analysis_result()
        excel_bytes = generate_excel_report(result, "ABC Sanayi")

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Özet"]

        # İlk satır başlık
        assert ws.cell(row=1, column=1).value == "Fiyatlama Analiz Raporu"
        # Dönem
        assert ws.cell(row=3, column=2).value == "2026-01"

    def test_simulation_sheet_has_rows(self):
        """Simülasyon sheet'inde satırlar var."""
        from openpyxl import load_workbook
        import io

        result = _sample_analysis_result()
        excel_bytes = generate_excel_report(result)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Simülasyon"]

        # Header + 3 veri satırı
        assert ws.cell(row=1, column=1).value == "Katsayı"
        assert ws.cell(row=2, column=1).value == 1.02
        assert ws.cell(row=4, column=1).value == 1.10

    def test_hourly_detail_sheet(self):
        """Saatlik detay sheet'inde veri var."""
        from openpyxl import load_workbook
        import io

        result = _sample_analysis_result()
        excel_bytes = generate_excel_report(result)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Saatlik Detay"]

        assert ws.cell(row=1, column=1).value == "Tarih"
        assert ws.cell(row=2, column=1).value == "2026-01-01"

    def test_loss_map_sheet(self):
        """Zarar haritası sheet'inde veriler var."""
        from openpyxl import load_workbook
        import io

        result = _sample_analysis_result()
        excel_bytes = generate_excel_report(result)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Zarar Haritası"]

        assert ws.cell(row=3, column=2).value == 218  # total_loss_hours
