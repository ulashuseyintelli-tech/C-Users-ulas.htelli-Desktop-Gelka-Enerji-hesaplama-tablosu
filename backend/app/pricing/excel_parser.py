"""
Pricing Risk Engine — EPİAŞ Excel & Tüketim Excel Ayrıştırıcı.

EPİAŞ uzlaştırma Excel dosyasını ve müşteri tüketim Excel dosyasını
ayrıştırarak yapılandırılmış veri modeline dönüştürür.

Desteklenen EPİAŞ Excel formatı (Uzlaştırma Dönemi Detayı):
- Sheet: "Uzlaştırma Dönemi Detayı" (veya ilk sheet)
- Sütun A: Tarih — datetime nesnesi (saat bilgisi gömülü)
- Sütun B: Versiyon — datetime (yoksayılır)
- Sütun C: Bölge — string "TR1" (filtre)
- Sütun D: PTF (TL / MWh) — int veya float
- Sütun E: SMF (TL / MWh) — int veya float
- Saat bilgisi Tarih sütunundaki datetime.hour'dan çıkarılır.

Requirements: 1.1, 1.2, 1.3, 1.4, 1.5, 2.4, 4.1, 4.3, 4.4, 18.1, 18.2, 18.3, 18.4
"""

from __future__ import annotations

import calendar
import io
import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

from openpyxl import load_workbook

from .models import ExcelParseResult, ConsumptionParseResult

logger = logging.getLogger(__name__)

# ═══════════════════════════════════════════════════════════════════════════════
# Sabitler
# ═══════════════════════════════════════════════════════════════════════════════

_PERIOD_RE = re.compile(r"^\d{4}-(0[1-9]|1[0-2])$")

# PTF/SMF değer aralığı
_PTF_MIN = 0.0
_PTF_MAX = 50_000.0
_SMF_MIN = 0.0
_SMF_MAX = 50_000.0

# Tüketim değer aralığı (negatif uyarı, reddetme yok)
_CONSUMPTION_MAX = 1_000_000.0  # kWh — makul üst sınır

# Header eşleştirme anahtar kelimeleri (case-insensitive, partial match)
_EPIAS_HEADER_KEYWORDS = {
    "tarih": "tarih",
    "date": "tarih",
    "ptf": "ptf",
    "piyasa_takas_fiyati": "ptf",
    "piyasa takas fiyatı": "ptf",
    "smf": "smf",
    "sistem_marjinal_fiyati": "smf",
    "sistem marjinal fiyatı": "smf",
    "bölge": "bolge",
    "bolge": "bolge",
    "region": "bolge",
}

_CONSUMPTION_HEADER_KEYWORDS = {
    "tarih": "tarih",
    "date": "tarih",
    "saat": "saat",
    "hour": "saat",
    "tüketim": "tuketim",
    "tuketim": "tuketim",
    "consumption": "tuketim",
    "kwh": "tuketim",
}


# ═══════════════════════════════════════════════════════════════════════════════
# Dahili veri yapıları — ayrıştırılmış kayıtlar
# ═══════════════════════════════════════════════════════════════════════════════


@dataclass
class ParsedMarketRecord:
    """Ayrıştırılmış tek saatlik piyasa verisi kaydı."""
    period: str       # YYYY-MM
    date: str         # YYYY-MM-DD
    hour: int         # 0-23
    ptf_tl_per_mwh: float
    smf_tl_per_mwh: float


@dataclass
class ParsedConsumptionRecord:
    """Ayrıştırılmış tek saatlik tüketim kaydı."""
    date: str         # YYYY-MM-DD
    hour: int         # 0-23
    consumption_kwh: float


@dataclass
class EpiasParseOutput:
    """parse_epias_excel() tam çıktısı — API result + ham kayıtlar."""
    result: ExcelParseResult
    records: list[ParsedMarketRecord] = field(default_factory=list)


@dataclass
class ConsumptionParseOutput:
    """parse_consumption_excel() tam çıktısı — API result + ham kayıtlar."""
    result: ConsumptionParseResult
    records: list[ParsedConsumptionRecord] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════════════════════
# Yardımcı fonksiyonlar
# ═══════════════════════════════════════════════════════════════════════════════


def expected_hours_for_period(period: str) -> int:
    """YYYY-MM formatında dönem için beklenen saat sayısı.

    Args:
        period: Dönem string'i, "YYYY-MM" formatında.

    Returns:
        Beklenen saat sayısı (672, 696, 720 veya 744).

    Raises:
        ValueError: Geçersiz dönem formatı.
    """
    if not _PERIOD_RE.match(period):
        raise ValueError(
            f"Geçersiz dönem formatı: '{period}'. Beklenen: YYYY-MM"
        )
    year = int(period[:4])
    month = int(period[5:7])
    days = calendar.monthrange(year, month)[1]
    return days * 24


def _safe_float(value: object) -> Optional[float]:
    """Hücre değerini float'a dönüştür. None dönerse geçersiz değer."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        cleaned = value.strip().replace(",", ".").replace(" ", "")
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _normalize_header(text: str) -> str:
    """Header metnini normalize et — küçük harf, boşluk/parantez temizle."""
    if not isinstance(text, str):
        return ""
    return text.strip().lower()


def _match_column(header_text: str, keywords: dict[str, str]) -> Optional[str]:
    """Header metnini anahtar kelimelerle eşleştir.

    Returns:
        Eşleşen mantıksal sütun adı (tarih, ptf, smf, bolge, saat, tuketim)
        veya None.
    """
    normalized = _normalize_header(header_text)
    if not normalized:
        return None

    for keyword, logical_name in keywords.items():
        if keyword in normalized:
            return logical_name
    return None


# ═══════════════════════════════════════════════════════════════════════════════
# EPİAŞ Excel Parser
# ═══════════════════════════════════════════════════════════════════════════════


def parse_epias_excel(
    file_bytes: bytes,
    filename: str,
) -> EpiasParseOutput:
    """EPİAŞ uzlaştırma Excel dosyasını ayrıştır.

    Gerçek EPİAŞ formatı:
    - Sheet: "Uzlaştırma Dönemi Detayı" (veya ilk sheet)
    - Tarih sütunu datetime nesnesi — saat bilgisi gömülü (hour 0-23)
    - Bölge sütunu "TR1" filtresi
    - PTF ve SMF sütunları int veya float

    Args:
        file_bytes: Excel dosyasının byte içeriği.
        filename: Orijinal dosya adı (loglama için).

    Returns:
        EpiasParseOutput: result (ExcelParseResult) + records listesi.
    """
    warnings: list[str] = []
    rejected_rows: list[dict] = []
    records: list[ParsedMarketRecord] = []

    # ── 1. Excel dosyasını aç ──────────────────────────────────────────────
    try:
        wb = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        logger.error("Excel dosyası açılamadı: %s — %s", filename, exc)
        return EpiasParseOutput(
            result=ExcelParseResult(
                success=False,
                period="",
                total_rows=0,
                expected_hours=0,
                missing_hours=[],
                rejected_rows=[{"row": 0, "reason": f"Excel dosyası açılamadı: {exc}"}],
                warnings=[],
                quality_score=0,
            ),
        )

    # ── 2. Sheet seç ──────────────────────────────────────────────────────
    target_sheet_names = [
        "uzlaştırma dönemi detayı",
        "uzlastirma donemi detayi",
        "uzlaştırma",
        "uzlastirma",
    ]
    ws = None
    for sn in wb.sheetnames:
        if sn.strip().lower() in target_sheet_names:
            ws = wb[sn]
            break
    if ws is None:
        # İlk sheet'i kullan
        ws = wb[wb.sheetnames[0]]
        warnings.append(
            f"'Uzlaştırma Dönemi Detayı' sheet'i bulunamadı, "
            f"ilk sheet kullanılıyor: '{wb.sheetnames[0]}'"
        )

    # ── 3. Header satırını bul ─────────────────────────────────────────────
    col_map: dict[str, int] = {}  # logical_name → column index (0-based)
    header_row_idx: Optional[int] = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
        temp_map: dict[str, int] = {}
        for cell in row:
            if cell.value is None:
                continue
            logical = _match_column(str(cell.value), _EPIAS_HEADER_KEYWORDS)
            if logical and logical not in temp_map:
                temp_map[logical] = cell.column - 1  # 0-based

        # En az Tarih ve PTF bulunmalı
        if "tarih" in temp_map and "ptf" in temp_map:
            col_map = temp_map
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        wb.close()
        return EpiasParseOutput(
            result=ExcelParseResult(
                success=False,
                period="",
                total_rows=0,
                expected_hours=0,
                missing_hours=[],
                rejected_rows=[{
                    "row": 0,
                    "reason": (
                        "EPİAŞ Excel formatı tanınamadı. "
                        "Beklenen sütunlar: Tarih, PTF (TL/MWh). "
                        "İlk 20 satırda header bulunamadı."
                    ),
                }],
                warnings=[],
                quality_score=0,
            ),
        )

    has_smf = "smf" in col_map
    has_bolge = "bolge" in col_map

    if not has_smf:
        warnings.append("SMF sütunu bulunamadı — SMF değerleri 0.0 olarak ayarlanacak.")

    # ── 4. Veri satırlarını ayrıştır ───────────────────────────────────────
    period_detected: Optional[str] = None
    seen_date_hours: dict[str, int] = {}  # "YYYY-MM-DD_HH" → count
    row_number = header_row_idx  # data starts after header

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row_number += 1

        # Boş satır kontrolü
        if row is None or all(c is None for c in row):
            continue

        # ── Tarih sütunu ──
        tarih_col = col_map["tarih"]
        tarih_val = row[tarih_col] if tarih_col < len(row) else None

        if tarih_val is None:
            continue  # Boş tarih — atla (genellikle dosya sonu)

        # Tarih ayrıştırma — datetime nesnesi veya string
        dt_obj: Optional[datetime] = None
        if isinstance(tarih_val, datetime):
            dt_obj = tarih_val
        elif isinstance(tarih_val, str):
            # DD.MM.YYYY veya YYYY-MM-DD formatı dene
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    dt_obj = datetime.strptime(tarih_val.strip(), fmt)
                    break
                except ValueError:
                    continue
            if dt_obj is None:
                rejected_rows.append({
                    "row": row_number,
                    "reason": f"Tarih ayrıştırılamadı: '{tarih_val}'",
                })
                continue
        else:
            rejected_rows.append({
                "row": row_number,
                "reason": f"Beklenmeyen tarih tipi: {type(tarih_val).__name__}",
            })
            continue

        date_str = dt_obj.strftime("%Y-%m-%d")
        hour = dt_obj.hour  # Saat bilgisi datetime'dan çıkarılır

        # Dönem çıkar
        row_period = dt_obj.strftime("%Y-%m")
        if period_detected is None:
            period_detected = row_period

        # ── Bölge filtresi ──
        if has_bolge:
            bolge_col = col_map["bolge"]
            bolge_val = row[bolge_col] if bolge_col < len(row) else None
            if bolge_val is not None:
                bolge_str = str(bolge_val).strip().upper()
                if bolge_str != "TR1":
                    continue  # TR1 olmayan bölgeleri atla

        # ── PTF ──
        ptf_col = col_map["ptf"]
        ptf_raw = row[ptf_col] if ptf_col < len(row) else None
        ptf_val = _safe_float(ptf_raw)

        if ptf_val is None:
            rejected_rows.append({
                "row": row_number,
                "reason": f"PTF değeri okunamadı: '{ptf_raw}'",
            })
            continue

        if ptf_val < _PTF_MIN or ptf_val > _PTF_MAX:
            rejected_rows.append({
                "row": row_number,
                "reason": (
                    f"PTF aralık dışı: {ptf_val:.2f} "
                    f"(beklenen: {_PTF_MIN}–{_PTF_MAX})"
                ),
            })
            continue

        # ── SMF ──
        smf_val = 0.0
        if has_smf:
            smf_col = col_map["smf"]
            smf_raw = row[smf_col] if smf_col < len(row) else None
            smf_parsed = _safe_float(smf_raw)

            if smf_parsed is None:
                rejected_rows.append({
                    "row": row_number,
                    "reason": f"SMF değeri okunamadı: '{smf_raw}'",
                })
                continue

            if smf_parsed < _SMF_MIN or smf_parsed > _SMF_MAX:
                rejected_rows.append({
                    "row": row_number,
                    "reason": (
                        f"SMF aralık dışı: {smf_parsed:.2f} "
                        f"(beklenen: {_SMF_MIN}–{_SMF_MAX})"
                    ),
                })
                continue

            smf_val = smf_parsed

        # ── Mükerrer saat kontrolü ──
        date_hour_key = f"{date_str}_{hour:02d}"
        seen_date_hours[date_hour_key] = seen_date_hours.get(date_hour_key, 0) + 1

        # Kayıt oluştur
        records.append(ParsedMarketRecord(
            period=row_period,
            date=date_str,
            hour=hour,
            ptf_tl_per_mwh=round(ptf_val, 2),
            smf_tl_per_mwh=round(smf_val, 2),
        ))

    wb.close()

    # ── 5. Dönem ve beklenen saat kontrolü ─────────────────────────────────
    if period_detected is None:
        return EpiasParseOutput(
            result=ExcelParseResult(
                success=False,
                period="",
                total_rows=0,
                expected_hours=0,
                missing_hours=[],
                rejected_rows=rejected_rows or [{
                    "row": 0,
                    "reason": "Hiç geçerli veri satırı bulunamadı.",
                }],
                warnings=warnings,
                quality_score=0,
            ),
        )

    exp_hours = expected_hours_for_period(period_detected)

    # ── 6. Mükerrer saat tespiti ───────────────────────────────────────────
    duplicate_keys: list[str] = []
    for key, count in seen_date_hours.items():
        if count > 1:
            duplicate_keys.append(key)

    if duplicate_keys:
        # Mükerrer kayıtları kaldır — sadece ilkini tut
        seen_once: set[str] = set()
        deduped_records: list[ParsedMarketRecord] = []
        for rec in records:
            key = f"{rec.date}_{rec.hour:02d}"
            if key not in seen_once:
                seen_once.add(key)
                deduped_records.append(rec)
        dup_count = len(records) - len(deduped_records)
        records = deduped_records
        warnings.append(
            f"{dup_count} mükerrer saat tespit edildi ve kaldırıldı: "
            f"{', '.join(sorted(duplicate_keys)[:10])}"
            + ("..." if len(duplicate_keys) > 10 else "")
        )

    # ── 7. Eksik saat tespiti ──────────────────────────────────────────────
    year = int(period_detected[:4])
    month = int(period_detected[5:7])
    days_in_month = calendar.monthrange(year, month)[1]

    expected_set: set[str] = set()
    for day in range(1, days_in_month + 1):
        for h in range(24):
            expected_set.add(f"{year:04d}-{month:02d}-{day:02d}_{h:02d}")

    actual_set = {f"{r.date}_{r.hour:02d}" for r in records}
    missing_keys = sorted(expected_set - actual_set)

    # missing_hours: sıralı indeks listesi (0-based, gün*24 + saat)
    missing_hours: list[int] = []
    for mk in missing_keys:
        parts = mk.split("_")
        day_part = int(parts[0].split("-")[2])
        hour_part = int(parts[1])
        idx = (day_part - 1) * 24 + hour_part
        missing_hours.append(idx)

    if missing_hours:
        warnings.append(
            f"{len(missing_hours)} eksik saat tespit edildi "
            f"(beklenen: {exp_hours}, mevcut: {len(records)})"
        )

    # ── 8. Kalite skoru hesaplama ──────────────────────────────────────────
    quality_score = _calculate_market_quality_score(
        total_valid=len(records),
        expected_hours=exp_hours,
        missing_count=len(missing_hours),
        rejected_count=len(rejected_rows),
        duplicate_count=len(duplicate_keys),
    )

    return EpiasParseOutput(
        result=ExcelParseResult(
            success=True,
            period=period_detected,
            total_rows=len(records),
            expected_hours=exp_hours,
            missing_hours=missing_hours,
            rejected_rows=rejected_rows,
            warnings=warnings,
            quality_score=quality_score,
        ),
        records=records,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Tüketim Excel Parser
# ═══════════════════════════════════════════════════════════════════════════════


def parse_consumption_excel(
    file_bytes: bytes,
    filename: str,
    customer_id: str,
) -> ConsumptionParseOutput:
    """Müşteri tüketim Excel dosyasını ayrıştır.

    İki format desteklenir:
    1. Tarih + Saat + Tüketim (kWh) — ayrı saat sütunu
    2. Tarih (datetime, saat gömülü) + Tüketim (kWh) — EPİAŞ benzeri format

    Args:
        file_bytes: Excel dosyasının byte içeriği.
        filename: Orijinal dosya adı (loglama için).
        customer_id: Müşteri kimliği.

    Returns:
        ConsumptionParseOutput: result (ConsumptionParseResult) + records listesi.
    """
    warnings: list[str] = []
    records: list[ParsedConsumptionRecord] = []
    negative_hours: list[int] = []

    # ── 1. Excel dosyasını aç ──────────────────────────────────────────────
    try:
        wb = load_workbook(
            filename=io.BytesIO(file_bytes),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        logger.error("Tüketim Excel dosyası açılamadı: %s — %s", filename, exc)
        return ConsumptionParseOutput(
            result=ConsumptionParseResult(
                success=False,
                customer_id=customer_id,
                period="",
                total_rows=0,
                total_kwh=0.0,
                negative_hours=[],
                quality_score=0,
            ),
        )

    # İlk sheet'i kullan
    ws = wb[wb.sheetnames[0]]

    # ── 2. Header satırını bul ─────────────────────────────────────────────
    col_map: dict[str, int] = {}
    header_row_idx: Optional[int] = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
        temp_map: dict[str, int] = {}
        for cell in row:
            if cell.value is None:
                continue
            logical = _match_column(str(cell.value), _CONSUMPTION_HEADER_KEYWORDS)
            if logical and logical not in temp_map:
                temp_map[logical] = cell.column - 1  # 0-based

        # En az Tarih ve Tüketim bulunmalı
        if "tarih" in temp_map and "tuketim" in temp_map:
            col_map = temp_map
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        wb.close()
        return ConsumptionParseOutput(
            result=ConsumptionParseResult(
                success=False,
                customer_id=customer_id,
                period="",
                total_rows=0,
                total_kwh=0.0,
                negative_hours=[],
                quality_score=0,
            ),
        )

    has_saat = "saat" in col_map

    # ── 3. Veri satırlarını ayrıştır ───────────────────────────────────────
    period_detected: Optional[str] = None
    total_kwh = 0.0
    row_number = header_row_idx
    record_index = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row_number += 1

        if row is None or all(c is None for c in row):
            continue

        # ── Tarih ──
        tarih_col = col_map["tarih"]
        tarih_val = row[tarih_col] if tarih_col < len(row) else None

        if tarih_val is None:
            continue

        dt_obj: Optional[datetime] = None
        if isinstance(tarih_val, datetime):
            dt_obj = tarih_val
        elif isinstance(tarih_val, str):
            for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    dt_obj = datetime.strptime(tarih_val.strip(), fmt)
                    break
                except ValueError:
                    continue
            if dt_obj is None:
                warnings.append(f"Satır {row_number}: Tarih ayrıştırılamadı: '{tarih_val}'")
                continue
        else:
            warnings.append(
                f"Satır {row_number}: Beklenmeyen tarih tipi: {type(tarih_val).__name__}"
            )
            continue

        date_str = dt_obj.strftime("%Y-%m-%d")

        # ── Saat ──
        if has_saat:
            saat_col = col_map["saat"]
            saat_val = row[saat_col] if saat_col < len(row) else None
            hour_parsed = _safe_float(saat_val)
            if hour_parsed is None or hour_parsed < 0 or hour_parsed > 23:
                warnings.append(
                    f"Satır {row_number}: Geçersiz saat değeri: '{saat_val}'"
                )
                continue
            hour = int(hour_parsed)
        else:
            # Saat bilgisi datetime'dan çıkarılır
            hour = dt_obj.hour

        # Dönem çıkar
        row_period = dt_obj.strftime("%Y-%m")
        if period_detected is None:
            period_detected = row_period

        # ── Tüketim ──
        tuketim_col = col_map["tuketim"]
        tuketim_raw = row[tuketim_col] if tuketim_col < len(row) else None
        tuketim_val = _safe_float(tuketim_raw)

        if tuketim_val is None:
            warnings.append(
                f"Satır {row_number}: Tüketim değeri okunamadı: '{tuketim_raw}'"
            )
            continue

        # Negatif tüketim — uyarı (reddetme yok)
        if tuketim_val < 0:
            negative_hours.append(record_index)
            warnings.append(
                f"Satır {row_number}: Negatif tüketim: {tuketim_val:.2f} kWh "
                f"(tarih: {date_str}, saat: {hour})"
            )

        total_kwh += tuketim_val

        records.append(ParsedConsumptionRecord(
            date=date_str,
            hour=hour,
            consumption_kwh=round(tuketim_val, 4),
        ))
        record_index += 1

    wb.close()

    # ── 4. Sonuç ──────────────────────────────────────────────────────────
    if period_detected is None:
        return ConsumptionParseOutput(
            result=ConsumptionParseResult(
                success=False,
                customer_id=customer_id,
                period="",
                total_rows=0,
                total_kwh=0.0,
                negative_hours=negative_hours,
                quality_score=0,
            ),
        )

    # Kalite skoru
    quality_score = _calculate_consumption_quality_score(
        total_valid=len(records),
        expected_hours=expected_hours_for_period(period_detected),
        negative_count=len(negative_hours),
    )

    return ConsumptionParseOutput(
        result=ConsumptionParseResult(
            success=True,
            customer_id=customer_id,
            period=period_detected,
            total_rows=len(records),
            total_kwh=round(total_kwh, 4),
            negative_hours=negative_hours,
            warnings=warnings,
            quality_score=quality_score,
        ),
        records=records,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Kalite Skoru Hesaplama
# ═══════════════════════════════════════════════════════════════════════════════


def _calculate_market_quality_score(
    total_valid: int,
    expected_hours: int,
    missing_count: int,
    rejected_count: int,
    duplicate_count: int,
) -> int:
    """Piyasa verisi kalite skoru hesapla (0–100).

    Formül:
        100 - missing_penalty - rejected_penalty - duplicate_penalty

    Cezalar:
        - Eksik saat: (missing / expected) × 50
        - Reddedilen satır: (rejected / max(expected, 1)) × 30
        - Mükerrer saat: (duplicate / max(expected, 1)) × 20
    """
    if expected_hours == 0:
        return 0 if total_valid == 0 else 100

    missing_penalty = (missing_count / expected_hours) * 50
    rejected_penalty = (rejected_count / max(expected_hours, 1)) * 30
    duplicate_penalty = (duplicate_count / max(expected_hours, 1)) * 20

    score = 100.0 - missing_penalty - rejected_penalty - duplicate_penalty
    return max(0, min(100, int(round(score))))


def _calculate_consumption_quality_score(
    total_valid: int,
    expected_hours: int,
    negative_count: int,
) -> int:
    """Tüketim verisi kalite skoru hesapla (0–100).

    Formül:
        100 - missing_penalty - negative_penalty

    Cezalar:
        - Eksik saat: (missing / expected) × 60
        - Negatif tüketim: (negative / max(total, 1)) × 40
    """
    if expected_hours == 0:
        return 0 if total_valid == 0 else 100

    missing_count = max(0, expected_hours - total_valid)
    missing_penalty = (missing_count / expected_hours) * 60
    negative_penalty = (negative_count / max(total_valid, 1)) * 40

    score = 100.0 - missing_penalty - negative_penalty
    return max(0, min(100, int(round(score))))
