"""
Pricing Risk Engine — Excel Dışa Aktarma (Formatter).

Saatlik piyasa verilerini ve tüketim profillerini EPİAŞ-uyumlu
Excel formatına geri yazar. Round-trip özelliği için kritik bileşen.

Requirements: 1.7, 1.8, 4.5, 4.6
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook

from .excel_parser import ParsedMarketRecord, ParsedConsumptionRecord


# ═══════════════════════════════════════════════════════════════════════════════
# Piyasa Verisi Excel Dışa Aktarma
# ═══════════════════════════════════════════════════════════════════════════════


def export_market_data_to_excel(
    records: list[ParsedMarketRecord],
    period: str,
) -> bytes:
    """Saatlik piyasa verilerini EPİAŞ-uyumlu Excel formatında dışa aktar.

    Sütunlar:
        - Tarih: datetime nesnesi (tarih + saat bilgisi gömülü)
        - Bölge: "TR1" (sabit)
        - PTF (TL/MWh): float
        - SMF (TL/MWh): float

    Args:
        records: Ayrıştırılmış piyasa verisi kayıtları.
        period: Dönem (YYYY-MM) — dosya adı/metadata için.

    Returns:
        Excel dosyasının byte içeriği (BytesIO).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Uzlaştırma Dönemi Detayı"

    # Header satırı
    ws.append(["Tarih", "Bölge", "PTF (TL/MWh)", "SMF (TL/MWh)"])

    # Veri satırları — tarih + saat bilgisini datetime olarak yaz
    for rec in records:
        dt = datetime.strptime(f"{rec.date} {rec.hour:02d}:00:00", "%Y-%m-%d %H:%M:%S")
        ws.append([dt, "TR1", rec.ptf_tl_per_mwh, rec.smf_tl_per_mwh])

    # Tarih sütunu formatı
    for row_idx in range(2, len(records) + 2):
        ws.cell(row=row_idx, column=1).number_format = "DD.MM.YYYY HH:MM"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ═══════════════════════════════════════════════════════════════════════════════
# Tüketim Verisi Excel Dışa Aktarma
# ═══════════════════════════════════════════════════════════════════════════════


def export_consumption_to_excel(
    records: list[ParsedConsumptionRecord],
    period: str,
) -> bytes:
    """Tüketim profilini Excel formatında dışa aktar.

    Sütunlar:
        - Tarih: tarih string'i (YYYY-MM-DD)
        - Saat: 0–23 integer
        - Tüketim (kWh): float

    Args:
        records: Ayrıştırılmış tüketim kayıtları.
        period: Dönem (YYYY-MM) — dosya adı/metadata için.

    Returns:
        Excel dosyasının byte içeriği (BytesIO).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tüketim Verisi"

    # Header satırı
    ws.append(["Tarih", "Saat", "Tüketim (kWh)"])

    # Veri satırları
    for rec in records:
        # Tarih'i datetime olarak yaz (parser her iki formatı da destekler)
        dt = datetime.strptime(rec.date, "%Y-%m-%d")
        ws.append([dt, rec.hour, rec.consumption_kwh])

    # Tarih sütunu formatı
    for row_idx in range(2, len(records) + 2):
        ws.cell(row=row_idx, column=1).number_format = "DD.MM.YYYY"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
