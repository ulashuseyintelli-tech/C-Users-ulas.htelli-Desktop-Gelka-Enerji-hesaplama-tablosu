"""
Pricing Risk Engine — PDF ve Excel Rapor Üretimi.

PDF: Jinja2 HTML template → Playwright/WeasyPrint/ReportLab fallback
Excel: openpyxl ile 5 sheet

Requirements: 17.1, 17.2, 17.3, 17.4
"""

from __future__ import annotations

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Optional

from jinja2 import Environment, FileSystemLoader

logger = logging.getLogger(__name__)

# Template dizini
_TEMPLATE_DIR = Path(__file__).parent.parent / "templates"


# ═══════════════════════════════════════════════════════════════════════════════
# PDF Rapor Üretimi
# ═══════════════════════════════════════════════════════════════════════════════


def generate_pdf_report(
    analysis_result: dict,
    customer_name: Optional[str] = None,
    contact_person: Optional[str] = None,
    report_mode: str = "internal",
) -> bytes:
    """Fiyatlama analiz raporu PDF üret.

    2 sayfa:
    - Sayfa 1: Özet + Fiyatlama + T1/T2/T3 dağılımı
    - Sayfa 2: Katsayı simülasyonu + Zarar haritası

    Args:
        analysis_result: /analyze endpoint çıktısı (dict).
        customer_name: Müşteri adı (opsiyonel).
        contact_person: İlgili kişi (opsiyonel).
        report_mode: "internal" (watermark yok) veya "demo" (watermark var).

    Returns:
        PDF bytes.
    """
    # Template context hazırla
    ctx = _build_template_context(analysis_result, customer_name, contact_person)
    ctx["report_mode"] = report_mode
    ctx["is_demo"] = report_mode == "demo"

    # HTML render
    env = Environment(
        loader=FileSystemLoader(str(_TEMPLATE_DIR)),
        autoescape=False,
    )
    template = env.get_template("pricing_analysis_template.html")
    html_content = template.render(**ctx)

    # PDF render — Playwright → WeasyPrint fallback
    pdf_bytes = _html_to_pdf(html_content)
    logger.info("Pricing PDF generated: %d bytes", len(pdf_bytes))
    return pdf_bytes


def _build_template_context(
    result: dict,
    customer_name: Optional[str],
    contact_person: Optional[str],
) -> dict:
    """Analiz sonucundan template context oluştur."""
    wp = result.get("weighted_prices", {})
    sc = result.get("supplier_cost", {})
    pricing = result.get("pricing", {})
    risk = result.get("risk_score", {})
    safe = result.get("safe_multiplier", {})
    loss = result.get("loss_map", {})
    tz = result.get("time_zone_breakdown", {})

    # Simülasyon verisi (varsa)
    simulation = result.get("simulation", [])

    # Risk reasons
    risk_reasons = risk.get("reasons", [])
    if not risk_reasons:
        risk_reasons = [f"Ağırlıklı PTF sapması: %{risk.get('deviation_pct', 0):.1f}"]

    # Coherence note
    coherence_note = None
    for w in result.get("warnings", []):
        if w.get("type") == "coherence_warning":
            coherence_note = w.get("message")
            break

    # Safe multiplier rounded (simülasyon tablosunda vurgulama için)
    safe_mult = safe.get("safe_multiplier", 1.1)
    import math
    safe_mult_rounded = math.ceil(safe_mult * 100) / 100

    return {
        "period": result.get("period", ""),
        "customer_name": customer_name or result.get("customer_id", ""),
        "contact_person": contact_person,
        "report_date": datetime.now().strftime("%d.%m.%Y"),

        # Fiyat verileri
        "weighted_ptf": wp.get("weighted_ptf_tl_per_mwh", 0),
        "yekdem": sc.get("yekdem_tl_per_mwh", 0),
        "imbalance_cost": sc.get("imbalance_tl_per_mwh", 0),
        "supplier_cost": sc.get("total_cost_tl_per_mwh", 0),
        "multiplier": pricing.get("multiplier", 1.0),
        "pricing": pricing,

        # Risk
        "risk_level": risk.get("score", "Düşük"),
        "deviation_pct": risk.get("deviation_pct", 0),
        "risk_reasons": risk_reasons,
        "coherence_note": coherence_note,

        # Güvenli katsayı
        "safe_multiplier": safe_mult,
        "recommended_multiplier": safe.get("recommended_multiplier", 1.10),
        "safe_multiplier_rounded": safe_mult_rounded,
        "safe_warning": safe.get("warning"),

        # Zaman dilimleri
        "time_zones": tz,

        # Simülasyon
        "simulation": simulation,

        # Zarar haritası
        "loss_map": loss,

        # Uyarılar
        "warnings": result.get("warnings", []),
    }


def _html_to_pdf(html_content: str) -> bytes:
    """HTML → PDF dönüşümü — Playwright → WeasyPrint fallback."""
    # 1. Playwright
    try:
        from ..services.pdf_playwright import html_to_pdf_bytes_sync_v2
        pdf_bytes = html_to_pdf_bytes_sync_v2(html_content)
        logger.info("Pricing PDF rendered with Playwright")
        return pdf_bytes
    except Exception as e:
        logger.warning("Playwright PDF failed: %s", e)

    # 2. WeasyPrint
    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html_content).write_pdf()
        logger.info("Pricing PDF rendered with WeasyPrint")
        return pdf_bytes
    except Exception as e:
        logger.warning("WeasyPrint PDF failed: %s", e)

    raise RuntimeError("PDF render failed: Playwright and WeasyPrint both unavailable")


# ═══════════════════════════════════════════════════════════════════════════════
# Excel Rapor Üretimi
# ═══════════════════════════════════════════════════════════════════════════════


def generate_excel_report(
    analysis_result: dict,
    customer_name: Optional[str] = None,
) -> bytes:
    """Fiyatlama analiz raporu Excel üret.

    5 Sheet:
    - Sheet 1: Özet
    - Sheet 2: T1/T2/T3 Dağılım
    - Sheet 3: Katsayı Simülasyonu
    - Sheet 4: Saatlik Detay (744 satır)
    - Sheet 5: Zarar Haritası

    Args:
        analysis_result: /analyze endpoint çıktısı (dict).
        customer_name: Müşteri adı (opsiyonel).

    Returns:
        Excel bytes (.xlsx).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # Stiller
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")
    num_fmt = '#,##0.00'
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    wp = analysis_result.get("weighted_prices", {})
    sc = analysis_result.get("supplier_cost", {})
    pricing = analysis_result.get("pricing", {})
    risk = analysis_result.get("risk_score", {})
    safe = analysis_result.get("safe_multiplier", {})
    tz = analysis_result.get("time_zone_breakdown", {})
    loss = analysis_result.get("loss_map", {})
    simulation = analysis_result.get("simulation", [])
    hour_costs = analysis_result.get("hour_costs", [])

    # ── Sheet 1: Özet ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Özet"
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25

    rows = [
        ("Fiyatlama Analiz Raporu", ""),
        ("", ""),
        ("Dönem", analysis_result.get("period", "")),
        ("Müşteri", customer_name or analysis_result.get("customer_id", "")),
        ("Rapor Tarihi", datetime.now().strftime("%d.%m.%Y")),
        ("", ""),
        ("— Tedarikçi Maliyet Yapısı —", ""),
        ("Ağırlıklı PTF", wp.get("weighted_ptf_tl_per_mwh", 0)),
        ("Aritmetik Ortalama PTF", wp.get("arithmetic_avg_ptf", 0)),
        ("YEKDEM Bedeli", sc.get("yekdem_tl_per_mwh", 0)),
        ("Dengesizlik Maliyeti", sc.get("imbalance_tl_per_mwh", 0)),
        ("Toplam Tedarikçi Maliyeti", sc.get("total_cost_tl_per_mwh", 0)),
        ("", ""),
        ("— Fiyatlama —", ""),
        ("Katsayı", pricing.get("multiplier", 0)),
        ("Satış Fiyatı (TL/MWh)", pricing.get("sales_price_tl_per_mwh", 0)),
        ("Brüt Marj (TL/MWh)", pricing.get("gross_margin_tl_per_mwh", 0)),
        ("Net Marj (TL/MWh)", pricing.get("net_margin_tl_per_mwh", 0)),
        ("Toplam Satış (TL)", pricing.get("total_sales_tl", 0)),
        ("Toplam Maliyet (TL)", pricing.get("total_cost_tl", 0)),
        ("Toplam Net Marj (TL)", pricing.get("total_net_margin_tl", 0)),
        ("", ""),
        ("— Risk —", ""),
        ("Risk Seviyesi", risk.get("score", "")),
        ("Sapma (%)", risk.get("deviation_pct", 0)),
        ("T2 Tüketim Payı (%)", risk.get("t2_consumption_pct", 0)),
        ("Güvenli Katsayı", safe.get("safe_multiplier", 0)),
        ("Önerilen Katsayı", safe.get("recommended_multiplier", 0)),
    ]

    for i, (label, value) in enumerate(rows, 1):
        ws1.cell(row=i, column=1, value=label)
        ws1.cell(row=i, column=2, value=value)
        if i == 1:
            ws1.cell(row=i, column=1).font = Font(bold=True, size=14)

    # ── Sheet 2: T1/T2/T3 Dağılım ─────────────────────────────────────
    ws2 = wb.create_sheet("T1-T2-T3 Dağılım")
    headers2 = ["Dilim", "Tüketim (kWh)", "Pay (%)", "Ağırlıklı PTF", "Ağırlıklı SMF", "Maliyet (TL)"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill

    row_idx = 2
    for tz_key in ["T1", "T2", "T3"]:
        tz_data = tz.get(tz_key, {})
        if isinstance(tz_data, dict):
            ws2.cell(row=row_idx, column=1, value=tz_data.get("label", tz_key))
            ws2.cell(row=row_idx, column=2, value=tz_data.get("consumption_kwh", 0))
            ws2.cell(row=row_idx, column=3, value=tz_data.get("consumption_pct", 0))
            ws2.cell(row=row_idx, column=4, value=tz_data.get("weighted_ptf_tl_per_mwh", 0))
            ws2.cell(row=row_idx, column=5, value=tz_data.get("weighted_smf_tl_per_mwh", 0))
            ws2.cell(row=row_idx, column=6, value=tz_data.get("total_cost_tl", 0))
            row_idx += 1

    # ── Sheet 3: Katsayı Simülasyonu ───────────────────────────────────
    ws3 = wb.create_sheet("Simülasyon")
    headers3 = ["Katsayı", "Satış (TL)", "Maliyet (TL)", "Brüt Marj", "Bayi Kom.", "Net Marj", "Zarar Saat", "Zarar TL"]
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill

    for i, row in enumerate(simulation, 2):
        if isinstance(row, dict):
            ws3.cell(row=i, column=1, value=row.get("multiplier", 0))
            ws3.cell(row=i, column=2, value=row.get("total_sales_tl", 0))
            ws3.cell(row=i, column=3, value=row.get("total_cost_tl", 0))
            ws3.cell(row=i, column=4, value=row.get("gross_margin_tl", 0))
            ws3.cell(row=i, column=5, value=row.get("dealer_commission_tl", 0))
            ws3.cell(row=i, column=6, value=row.get("net_margin_tl", 0))
            ws3.cell(row=i, column=7, value=row.get("loss_hours", 0))
            ws3.cell(row=i, column=8, value=row.get("total_loss_tl", 0))

    # ── Sheet 4: Saatlik Detay ─────────────────────────────────────────
    ws4 = wb.create_sheet("Saatlik Detay")
    headers4 = ["Tarih", "Saat", "Tüketim (kWh)", "PTF (TL/MWh)", "SMF (TL/MWh)",
                "YEKDEM", "Baz Maliyet (TL)", "Satış (TL)", "Marj (TL)", "Zarar?", "Dilim"]
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill

    for i, hc in enumerate(hour_costs, 2):
        if isinstance(hc, dict):
            ws4.cell(row=i, column=1, value=hc.get("date", ""))
            ws4.cell(row=i, column=2, value=hc.get("hour", 0))
            ws4.cell(row=i, column=3, value=hc.get("consumption_kwh", 0))
            ws4.cell(row=i, column=4, value=hc.get("ptf_tl_per_mwh", 0))
            ws4.cell(row=i, column=5, value=hc.get("smf_tl_per_mwh", 0))
            ws4.cell(row=i, column=6, value=hc.get("yekdem_tl_per_mwh", 0))
            ws4.cell(row=i, column=7, value=hc.get("base_cost_tl", 0))
            ws4.cell(row=i, column=8, value=hc.get("sales_price_tl", 0))
            ws4.cell(row=i, column=9, value=hc.get("margin_tl", 0))
            ws4.cell(row=i, column=10, value="Evet" if hc.get("is_loss_hour") else "")
            ws4.cell(row=i, column=11, value=hc.get("time_zone", ""))

    # ── Sheet 5: Zarar Haritası ────────────────────────────────────────
    ws5 = wb.create_sheet("Zarar Haritası")
    ws5.cell(row=1, column=1, value="Zarar Haritası Özeti").font = Font(bold=True, size=12)
    ws5.cell(row=3, column=1, value="Toplam Zararlı Saat")
    ws5.cell(row=3, column=2, value=loss.get("total_loss_hours", 0))
    ws5.cell(row=4, column=1, value="Toplam Zarar (TL)")
    ws5.cell(row=4, column=2, value=loss.get("total_loss_tl", 0))
    ws5.cell(row=5, column=1, value="T1 Zarar Saati")
    ws5.cell(row=5, column=2, value=loss.get("by_time_zone", {}).get("T1", 0))
    ws5.cell(row=6, column=1, value="T2 Zarar Saati")
    ws5.cell(row=6, column=2, value=loss.get("by_time_zone", {}).get("T2", 0))
    ws5.cell(row=7, column=1, value="T3 Zarar Saati")
    ws5.cell(row=7, column=2, value=loss.get("by_time_zone", {}).get("T3", 0))

    worst = loss.get("worst_hours", [])
    if worst:
        ws5.cell(row=9, column=1, value="En Kötü Saatler").font = Font(bold=True)
        wh_headers = ["Tarih", "Saat", "PTF", "Satış", "Zarar (TL)"]
        for col, h in enumerate(wh_headers, 1):
            cell = ws5.cell(row=10, column=col, value=h)
            cell.font = header_font
        for i, wh in enumerate(worst[:20], 11):
            ws5.cell(row=i, column=1, value=wh.get("date", ""))
            ws5.cell(row=i, column=2, value=wh.get("hour", 0))
            ws5.cell(row=i, column=3, value=wh.get("ptf", 0))
            ws5.cell(row=i, column=4, value=wh.get("sales_price", 0))
            ws5.cell(row=i, column=5, value=wh.get("loss_tl", 0))

    # Bytes'a dönüştür
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.getvalue()
    logger.info("Pricing Excel generated: %d bytes, %d sheets", len(excel_bytes), len(wb.sheetnames))
    return excel_bytes
