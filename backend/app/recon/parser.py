"""
Invoice Reconciliation Engine — Excel Parser.

IC-1: Tüm kWh değerleri Decimal olarak parse edilir.
IC-2: Timestamp'lar Europe/Istanbul'a normalize edilir.
IC-5: Pluggable provider format mimarisi (BaseFormatProvider + registry).

Desteklenen formatlar:
- Format A (büyük tüketici): "Profil Tarihi" + "Tüketim (Çekiş)" + "Çarpan"
- Format B (küçük tüketici): "Tarih" + "Aktif Çekiş"
"""

from __future__ import annotations

import re
from abc import ABC, abstractmethod
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .schemas import ExcelFormat, HourlyRecord, ParseError, ParseResult

# ═══════════════════════════════════════════════════════════════════════════════
# Constants
# ═══════════════════════════════════════════════════════════════════════════════

ISTANBUL_TZ = ZoneInfo("Europe/Istanbul")
MAX_FILE_SIZE_BYTES = 50 * 1024 * 1024  # 50 MB
MAX_HEADER_SCAN_ROWS = 10

# Date regex: DD/MM/YYYY HH:MM:SS
DATE_REGEX = re.compile(
    r"^(\d{1,2})/(\d{1,2})/(\d{4})\s+(\d{1,2}):(\d{2}):(\d{2})$"
)


# ═══════════════════════════════════════════════════════════════════════════════
# Exceptions
# ═══════════════════════════════════════════════════════════════════════════════


class ParserError(Exception):
    """Base parser error."""
    pass


class UnknownFormatError(ParserError):
    """Kolon başlıkları bilinen hiçbir formata uymuyor."""
    pass


class EmptyFileError(ParserError):
    """Dosya boş veya tüketim verisi bulunamadı."""
    pass


class FileTooLargeError(ParserError):
    """Dosya boyutu 50 MB limitini aşıyor."""
    pass


# ═══════════════════════════════════════════════════════════════════════════════
# IC-5: Pluggable Provider Architecture
# ═══════════════════════════════════════════════════════════════════════════════


class BaseFormatProvider(ABC):
    """Abstract base class for Excel format providers.

    IC-5: Yeni provider eklemek core parser mantığını değiştirmez.
    Sadece bu class'tan türetip PROVIDER_REGISTRY'ye eklemek yeterli.
    """

    @classmethod
    @abstractmethod
    def detect(cls, headers: list[str]) -> bool:
        """Kolon başlıklarından bu formatı algıla."""
        ...

    @classmethod
    @abstractmethod
    def get_format(cls) -> ExcelFormat:
        """Bu provider'ın ExcelFormat enum değeri."""
        ...

    @classmethod
    @abstractmethod
    def parse_sheet(
        cls, sheet: Worksheet, header_row: int, col_map: dict[str, int]
    ) -> tuple[list[HourlyRecord], list[ParseError], list[str], Optional[Decimal]]:
        """Sheet'i parse et.

        Returns:
            (records, errors, warnings, multiplier_metadata)
        """
        ...


class FormatAProvider(BaseFormatProvider):
    """Büyük tüketici formatı: Profil Tarihi + Tüketim (Çekiş) + Çarpan."""

    REQUIRED_COLS = {"profil tarihi", "tüketim (çekiş)"}
    OPTIONAL_COLS = {"çarpan"}

    @classmethod
    def detect(cls, headers: list[str]) -> bool:
        normalized = {h.strip().lower() for h in headers if h}
        return cls.REQUIRED_COLS.issubset(normalized)

    @classmethod
    def get_format(cls) -> ExcelFormat:
        return ExcelFormat.FORMAT_A

    @classmethod
    def parse_sheet(
        cls, sheet: Worksheet, header_row: int, col_map: dict[str, int]
    ) -> tuple[list[HourlyRecord], list[ParseError], list[str], Optional[Decimal]]:
        records: list[HourlyRecord] = []
        errors: list[ParseError] = []
        warnings: list[str] = []
        multiplier_meta: Optional[Decimal] = None

        date_col = col_map["profil tarihi"]
        consumption_col = col_map["tüketim (çekiş)"]
        multiplier_col = col_map.get("çarpan")

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=False),
            start=header_row + 1,
        ):
            # Skip empty rows
            date_cell = row[date_col].value
            consumption_cell = row[consumption_col].value
            if date_cell is None and consumption_cell is None:
                continue

            # Parse datetime
            ts = _parse_datetime(date_cell)
            if ts is None:
                errors.append(ParseError(
                    row_number=row_idx,
                    column="Profil Tarihi",
                    raw_value=str(date_cell) if date_cell else "",
                    error="Tarih parse edilemedi (beklenen: DD/MM/YYYY HH:MM:SS)",
                ))
                continue

            # Parse kWh
            kwh = _parse_kwh_value(consumption_cell)
            if kwh is None:
                errors.append(ParseError(
                    row_number=row_idx,
                    column="Tüketim (Çekiş)",
                    raw_value=str(consumption_cell) if consumption_cell else "",
                    error="kWh değeri parse edilemedi",
                ))
                continue

            # Handle negative consumption
            if kwh < Decimal("0"):
                warnings.append(
                    f"Satır {row_idx}: Negatif tüketim ({kwh}), mutlak değer kullanılıyor"
                )
                kwh = abs(kwh)

            # Parse multiplier (metadata only — NEVER applied)
            multiplier: Optional[Decimal] = None
            if multiplier_col is not None:
                mult_cell = row[multiplier_col].value
                if mult_cell is not None:
                    multiplier = _parse_kwh_value(mult_cell)
                    if multiplier is not None and multiplier_meta is None:
                        multiplier_meta = multiplier

            records.append(HourlyRecord(
                timestamp=ts,
                date=ts.strftime("%Y-%m-%d"),
                hour=ts.hour,
                period=ts.strftime("%Y-%m"),
                consumption_kwh=kwh,
                multiplier=multiplier,
            ))

        return records, errors, warnings, multiplier_meta


class FormatBProvider(BaseFormatProvider):
    """Küçük tüketici formatı: Tarih + Aktif Çekiş."""

    REQUIRED_COLS = {"tarih", "aktif çekiş"}

    @classmethod
    def detect(cls, headers: list[str]) -> bool:
        normalized = {h.strip().lower() for h in headers if h}
        return cls.REQUIRED_COLS.issubset(normalized)

    @classmethod
    def get_format(cls) -> ExcelFormat:
        return ExcelFormat.FORMAT_B

    @classmethod
    def parse_sheet(
        cls, sheet: Worksheet, header_row: int, col_map: dict[str, int]
    ) -> tuple[list[HourlyRecord], list[ParseError], list[str], Optional[Decimal]]:
        records: list[HourlyRecord] = []
        errors: list[ParseError] = []
        warnings: list[str] = []

        date_col = col_map["tarih"]
        consumption_col = col_map["aktif çekiş"]

        for row_idx, row in enumerate(
            sheet.iter_rows(min_row=header_row + 1, values_only=False),
            start=header_row + 1,
        ):
            date_cell = row[date_col].value
            consumption_cell = row[consumption_col].value
            if date_cell is None and consumption_cell is None:
                continue

            ts = _parse_datetime(date_cell)
            if ts is None:
                errors.append(ParseError(
                    row_number=row_idx,
                    column="Tarih",
                    raw_value=str(date_cell) if date_cell else "",
                    error="Tarih parse edilemedi (beklenen: DD/MM/YYYY HH:MM:SS)",
                ))
                continue

            kwh = _parse_kwh_value(consumption_cell)
            if kwh is None:
                errors.append(ParseError(
                    row_number=row_idx,
                    column="Aktif Çekiş",
                    raw_value=str(consumption_cell) if consumption_cell else "",
                    error="kWh değeri parse edilemedi",
                ))
                continue

            if kwh < Decimal("0"):
                warnings.append(
                    f"Satır {row_idx}: Negatif tüketim ({kwh}), mutlak değer kullanılıyor"
                )
                kwh = abs(kwh)

            records.append(HourlyRecord(
                timestamp=ts,
                date=ts.strftime("%Y-%m-%d"),
                hour=ts.hour,
                period=ts.strftime("%Y-%m"),
                consumption_kwh=kwh,
                multiplier=None,
            ))

        return records, errors, warnings, None


# ═══════════════════════════════════════════════════════════════════════════════
# Provider Registry (IC-5)
# ═══════════════════════════════════════════════════════════════════════════════

PROVIDER_REGISTRY: dict[str, type[BaseFormatProvider]] = {
    "provider_format_a": FormatAProvider,
    "provider_format_b": FormatBProvider,
}


# ═══════════════════════════════════════════════════════════════════════════════
# Core Parse Functions
# ═══════════════════════════════════════════════════════════════════════════════


def parse_excel(file_bytes: bytes) -> ParseResult:
    """Ana parse fonksiyonu — format algıla ve uygun provider'ı çağır.

    Args:
        file_bytes: Excel dosyası byte içeriği

    Returns:
        ParseResult with records, errors, warnings

    Raises:
        FileTooLargeError: Dosya > 50 MB
        EmptyFileError: Dosya boş veya veri yok
        UnknownFormatError: Tanınmayan format
    """
    # File size check
    if len(file_bytes) > MAX_FILE_SIZE_BYTES:
        raise FileTooLargeError(
            f"Dosya boyutu ({len(file_bytes) / (1024*1024):.1f} MB) "
            f"50 MB limitini aşıyor."
        )

    if len(file_bytes) == 0:
        raise EmptyFileError("Dosya boş veya tüketim verisi bulunamadı")

    # Load workbook
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = _find_data_sheet(wb)

    if sheet is None:
        raise EmptyFileError("Dosya boş veya tüketim verisi bulunamadı")

    # Detect format
    header_row, col_map, provider = _detect_format(sheet)

    if provider is None:
        raise UnknownFormatError(
            "Tanınmayan Excel formatı. Beklenen kolonlar: "
            "[Format A: 'Profil Tarihi' + 'Tüketim (Çekiş)'] veya "
            "[Format B: 'Tarih' + 'Aktif Çekiş']"
        )

    # Parse with detected provider
    records, errors, warnings, multiplier_meta = provider.parse_sheet(
        sheet, header_row, col_map
    )

    # Sort chronologically (IC-2: normalize to Istanbul TZ already done in _parse_datetime)
    if records:
        original_order = [r.timestamp for r in records]
        records.sort(key=lambda r: r.timestamp)
        sorted_order = [r.timestamp for r in records]
        if original_order != sorted_order:
            warnings.append("Kayıtlar kronolojik sıraya göre yeniden sıralandı")

    total_rows = len(records) + len(errors)
    success = len(records) > 0

    if not success and total_rows == 0:
        raise EmptyFileError("Dosya boş veya tüketim verisi bulunamadı")

    wb.close()

    return ParseResult(
        success=success,
        format_detected=provider.get_format(),
        records=records,
        errors=errors,
        total_rows=total_rows,
        successful_rows=len(records),
        failed_rows=len(errors),
        warnings=warnings,
        multiplier_metadata=multiplier_meta,
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Internal Helpers
# ═══════════════════════════════════════════════════════════════════════════════


def _find_data_sheet(wb) -> Optional[Worksheet]:
    """İlk sheet'i döndür veya tüketim verisi içeren sheet'i bul."""
    if not wb.sheetnames:
        return None
    # Default: first sheet
    return wb[wb.sheetnames[0]]


def _detect_format(
    sheet: Worksheet,
) -> tuple[int, dict[str, int], Optional[type[BaseFormatProvider]]]:
    """Kolon başlıklarını tarayarak format ve kolon indekslerini belirle.

    Returns:
        (header_row_index, col_map, provider_class) or (0, {}, None) if not found
    """
    try:
        max_row = sheet.max_row or 0
    except Exception:
        max_row = 0

    if max_row == 0:
        return 0, {}, None

    scan_limit = min(MAX_HEADER_SCAN_ROWS, max_row)

    for row_idx in range(1, scan_limit + 1):
        row_values = []
        try:
            for cell in sheet[row_idx]:
                row_values.append(str(cell.value).strip() if cell.value else "")
        except (IndexError, TypeError):
            continue

        if not any(row_values):
            continue

        # Check each provider in registry
        for provider_cls in PROVIDER_REGISTRY.values():
            if provider_cls.detect(row_values):
                # Build column map
                col_map: dict[str, int] = {}
                for col_idx, val in enumerate(row_values):
                    if val:
                        col_map[val.lower()] = col_idx
                return row_idx, col_map, provider_cls

    return 0, {}, None


def _parse_datetime(value: Any) -> Optional[datetime]:
    """Parse datetime from Excel cell value.

    IC-2: Sonuç Europe/Istanbul timezone'una normalize edilir.

    Supports:
    - DD/MM/YYYY HH:MM:SS string format
    - Native Excel datetime objects (openpyxl)
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        # Native datetime — assume Istanbul if naive
        if value.tzinfo is None:
            return value.replace(tzinfo=ISTANBUL_TZ)
        return value.astimezone(ISTANBUL_TZ)

    # String parse
    text = str(value).strip()
    if not text:
        return None

    match = DATE_REGEX.match(text)
    if match:
        day, month, year, hour, minute, second = (int(g) for g in match.groups())
        try:
            dt = datetime(year, month, day, hour, minute, second, tzinfo=ISTANBUL_TZ)
            # Validate hour range
            if not (0 <= dt.hour <= 23):
                return None
            return dt
        except (ValueError, OverflowError):
            return None

    return None


def _parse_kwh_value(value: Any) -> Optional[Decimal]:
    """Parse kWh value from Excel cell.

    IC-1: Sonuç Decimal olarak döner.

    Supports:
    - Numeric values (int, float)
    - Turkish locale strings: "1.234,56" → 1234.56
    - Plain strings: "1234.56"
    """
    if value is None:
        return None

    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    text = str(value).strip()
    if not text:
        return None

    # Remove whitespace (thousands separator)
    text = text.replace(" ", "")

    # Turkish locale detection: if both dot and comma present,
    # dot is thousands separator, comma is decimal
    if "," in text and "." in text:
        # "1.234,56" → "1234.56"
        text = text.replace(".", "").replace(",", ".")
    elif "," in text and "." not in text:
        # "1234,56" → "1234.56" (comma as decimal)
        text = text.replace(",", ".")
    # else: "1234.56" stays as is (dot as decimal)

    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None
